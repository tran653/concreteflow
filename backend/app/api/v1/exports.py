"""
API endpoints pour l'export de documents (PDF, Excel).
"""
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from io import BytesIO

from app.core.database import get_db
from app.models.user import User
from app.models.projet import Projet
from app.models.calcul import Calcul, CalculStatus
from app.models.element import Element
from app.api.deps import get_current_active_user
from app.services.export import generate_note_calcul, ExcelGenerator, generate_plan_de_pose_from_calcul
from app.models.calcul import TypeProduit

router = APIRouter(prefix="/exports", tags=["Exports"])


@router.get("/calcul/{calcul_id}/pdf")
async def export_calcul_pdf(
    calcul_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Exporte une note de calcul en PDF.
    """
    # Récupérer le calcul
    result = await db.execute(
        select(Calcul).join(Projet).where(
            Calcul.id == calcul_id,
            Projet.tenant_id == current_user.tenant_id
        )
    )
    calcul = result.scalar_one_or_none()

    if not calcul:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Calcul non trouvé"
        )

    if calcul.status != CalculStatus.COMPLETED:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Le calcul doit être terminé pour être exporté"
        )

    # Récupérer le projet
    result = await db.execute(
        select(Projet).where(Projet.id == calcul.projet_id)
    )
    projet = result.scalar_one_or_none()

    # Générer le PDF
    calcul_dict = {
        "name": calcul.name,
        "type_produit": calcul.type_produit.value,
        "norme": calcul.norme,
        "parametres": calcul.parametres,
        "resultats": calcul.resultats
    }

    projet_dict = {
        "name": projet.name,
        "reference": projet.reference,
        "client_name": projet.client_name
    } if projet else None

    pdf_content = generate_note_calcul(calcul_dict, projet_dict)

    # Retourner le fichier
    filename = f"note_calcul_{calcul.name.replace(' ', '_')}.pdf"

    return StreamingResponse(
        BytesIO(pdf_content),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


@router.get("/calcul/{calcul_id}/plan-de-pose")
async def export_plan_de_pose(
    calcul_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Exporte le plan de pose en PDF pour un plancher poutrelles-hourdis.

    Le plan de pose montre:
    - La disposition des poutrelles
    - L'espacement (entraxe)
    - Les dimensions du plancher
    - Les informations techniques
    """
    # Récupérer le calcul
    result = await db.execute(
        select(Calcul).join(Projet).where(
            Calcul.id == calcul_id,
            Projet.tenant_id == current_user.tenant_id
        )
    )
    calcul = result.scalar_one_or_none()

    if not calcul:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Calcul non trouvé"
        )

    if calcul.status != CalculStatus.COMPLETED:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Le calcul doit être terminé pour exporter le plan de pose"
        )

    if calcul.type_produit != TypeProduit.PLANCHER_POUTRELLES_HOURDIS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Le plan de pose n'est disponible que pour les planchers poutrelles-hourdis"
        )

    # Récupérer le projet
    result = await db.execute(
        select(Projet).where(Projet.id == calcul.projet_id)
    )
    projet = result.scalar_one_or_none()

    # Préparer les données
    calcul_dict = {
        "name": calcul.name,
        "computed_at": calcul.computed_at.isoformat() if calcul.computed_at else None,
        "parametres": calcul.parametres,
        "resultats": calcul.resultats
    }

    projet_dict = {
        "name": projet.name if projet else "N/A",
        "reference": projet.reference if projet else "N/A",
        "client_name": projet.client_name if projet else "N/A"
    }

    # Générer le PDF
    pdf_buffer = generate_plan_de_pose_from_calcul(calcul_dict, projet_dict)

    # Retourner le fichier
    filename = f"plan_de_pose_{calcul.name.replace(' ', '_')}.pdf"

    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


@router.get("/projet/{projet_id}/nomenclature")
async def export_nomenclature(
    projet_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Exporte la nomenclature des éléments d'un projet en Excel.
    """
    # Vérifier le projet
    result = await db.execute(
        select(Projet).where(
            Projet.id == projet_id,
            Projet.tenant_id == current_user.tenant_id
        )
    )
    projet = result.scalar_one_or_none()

    if not projet:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Projet non trouvé"
        )

    # Récupérer les éléments
    result = await db.execute(
        select(Element).join(Calcul).where(
            Calcul.projet_id == projet_id
        )
    )
    elements_db = result.scalars().all()

    # Convertir en dictionnaires
    elements = []
    for elem in elements_db:
        elements.append({
            "reference": elem.reference,
            "type": elem.calcul.type_produit.value if elem.calcul else "autre",
            "dimensions": {
                "length_mm": elem.longueur or 0,
                "width_mm": elem.largeur or 0,
                "height_mm": elem.hauteur or 0
            },
            "classe_beton": elem.classe_beton or "C30/37",
            "ferraillage": elem.ferraillage or {},
            "poids_kg": elem.poids or 0
        })

    # Si pas d'éléments, créer des données depuis les calculs
    if not elements:
        result = await db.execute(
            select(Calcul).where(Calcul.projet_id == projet_id)
        )
        calculs = result.scalars().all()

        for i, calcul in enumerate(calculs):
            geom = calcul.parametres.get('geometrie', {})
            elements.append({
                "reference": f"{calcul.type_produit.value[0].upper()}{i+1:03d}",
                "type": calcul.type_produit.value,
                "dimensions": {
                    "length_mm": geom.get('portee', 0) * 1000,
                    "width_mm": geom.get('largeur', 0) * 1000,
                    "height_mm": geom.get('hauteur', 0) * 1000
                },
                "classe_beton": calcul.parametres.get('materiaux', {}).get('classe_beton', 'C30/37'),
                "ferraillage": calcul.resultats.get('ferraillage', {}),
            })

    # Générer l'Excel
    generator = ExcelGenerator()
    projet_dict = {
        "name": projet.name,
        "reference": projet.reference,
        "client_name": projet.client_name
    }

    excel_content = generator.generate_nomenclature(elements, projet_dict)

    filename = f"nomenclature_{projet.reference}.xlsx"

    return StreamingResponse(
        BytesIO(excel_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


@router.get("/projet/{projet_id}/quantitatif")
async def export_quantitatif(
    projet_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Exporte le quantitatif matériaux d'un projet en Excel.
    """
    # Vérifier le projet
    result = await db.execute(
        select(Projet).where(
            Projet.id == projet_id,
            Projet.tenant_id == current_user.tenant_id
        )
    )
    projet = result.scalar_one_or_none()

    if not projet:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Projet non trouvé"
        )

    # Récupérer les calculs
    result = await db.execute(
        select(Calcul).where(
            Calcul.projet_id == projet_id,
            Calcul.status == CalculStatus.COMPLETED
        )
    )
    calculs_db = result.scalars().all()

    calculs = [
        {
            "name": c.name,
            "type_produit": c.type_produit.value,
            "parametres": c.parametres,
            "resultats": c.resultats
        }
        for c in calculs_db
    ]

    # Générer l'Excel
    generator = ExcelGenerator()
    projet_dict = {
        "name": projet.name,
        "reference": projet.reference,
        "client_name": projet.client_name
    }

    excel_content = generator.generate_quantitatif(calculs, projet_dict)

    filename = f"quantitatif_{projet.reference}.xlsx"

    return StreamingResponse(
        BytesIO(excel_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


@router.get("/calcul/{calcul_id}/excel")
async def export_calcul_excel(
    calcul_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Exporte les résultats d'un calcul en Excel.
    """
    # Récupérer le calcul
    result = await db.execute(
        select(Calcul).join(Projet).where(
            Calcul.id == calcul_id,
            Projet.tenant_id == current_user.tenant_id
        )
    )
    calcul = result.scalar_one_or_none()

    if not calcul:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Calcul non trouvé"
        )

    # Créer un Excel simple avec les résultats
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Résultats"

    # En-tête
    ws['A1'] = f"Calcul: {calcul.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Type: {calcul.type_produit.value}"
    ws['A3'] = f"Norme: {calcul.norme}"

    row = 5

    # Paramètres
    ws[f'A{row}'] = "PARAMÈTRES"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    for section, values in calcul.parametres.items():
        ws[f'A{row}'] = section.upper()
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        if isinstance(values, dict):
            for key, value in values.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = str(value)
                row += 1
        row += 1

    # Résultats
    ws[f'A{row}'] = "RÉSULTATS"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    for section, values in calcul.resultats.items():
        ws[f'A{row}'] = section.upper()
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        if isinstance(values, dict):
            for key, value in values.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = str(value)
                row += 1
        row += 1

    # Sauvegarder
    buffer = BytesIO()
    wb.save(buffer)
    excel_content = buffer.getvalue()
    buffer.close()

    filename = f"calcul_{calcul.name.replace(' ', '_')}.xlsx"

    return StreamingResponse(
        BytesIO(excel_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )
