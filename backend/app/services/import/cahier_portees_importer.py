"""
Service d'import des cahiers de portées limites depuis Excel/CSV.
Supporte différents formats fabricants avec détection automatique.
"""
from openpyxl import load_workbook
from typing import Dict, Any, List, Optional, Tuple
from io import BytesIO
from dataclasses import dataclass, field
import re


@dataclass
class ImportResult:
    """Résultat d'import."""
    success: bool
    lignes_importees: int
    lignes_ignorees: int
    erreurs: List[str] = field(default_factory=list)
    avertissements: List[str] = field(default_factory=list)


class CahierPorteesImporter:
    """Importateur de cahiers de portées depuis Excel."""

    # Charges standards attendues (kg/m² ou daN/m²)
    CHARGES_STANDARD = [250, 300, 350, 400, 450, 500, 550, 600, 650, 700, 750]

    # Patterns de détection colonnes
    PATTERNS_COLONNES = {
        'reference': r'(ref|réf|référence|poutrelle|type|désignation)',
        'hourdis': r'(hourdis|hauteur.*hourdis|h.*hourdis|h\s*=)',
        'entraxe': r'(entraxe|e/e|espacement|e\s*=)',
        'table': r'(table|compression|béton.*coulé|dc)',
    }

    def __init__(self):
        self.erreurs: List[str] = []
        self.avertissements: List[str] = []

    def import_from_excel(
        self,
        file_content: bytes,
        mapping_colonnes: Optional[Dict[str, int]] = None
    ) -> Tuple[List[Dict[str, Any]], ImportResult]:
        """
        Importe un cahier de portées depuis un fichier Excel.

        Args:
            file_content: Contenu binaire du fichier Excel
            mapping_colonnes: Mapping manuel colonnes (optionnel)

        Returns:
            Tuple (lignes_data, ImportResult)
        """
        self.erreurs = []
        self.avertissements = []
        lignes_data = []

        try:
            wb = load_workbook(filename=BytesIO(file_content), data_only=True)
            ws = wb.active

            # Détection automatique du mapping si non fourni
            if not mapping_colonnes:
                mapping_colonnes = self._detect_columns(ws)

            if not mapping_colonnes:
                self.erreurs.append("Impossible de détecter la structure du fichier")
                return [], ImportResult(False, 0, 0, self.erreurs, self.avertissements)

            # Détection ligne d'en-tête et charges
            header_row, charges = self._detect_header_and_charges(ws, mapping_colonnes)

            if not charges:
                self.erreurs.append("Aucune colonne de charge détectée")
                return [], ImportResult(False, 0, 0, self.erreurs, self.avertissements)

            self.avertissements.append(f"Charges détectées: {sorted(charges.values())}")

            # Import des lignes
            for row_idx in range(header_row + 1, ws.max_row + 1):
                try:
                    ligne = self._parse_row(ws, row_idx, mapping_colonnes, charges)
                    if ligne:
                        lignes_data.append(ligne)
                except Exception as e:
                    self.avertissements.append(f"Ligne {row_idx}: {str(e)}")

            return lignes_data, ImportResult(
                success=len(lignes_data) > 0,
                lignes_importees=len(lignes_data),
                lignes_ignorees=ws.max_row - header_row - len(lignes_data),
                erreurs=self.erreurs,
                avertissements=self.avertissements
            )

        except Exception as e:
            self.erreurs.append(f"Erreur lecture fichier: {str(e)}")
            return [], ImportResult(False, 0, 0, self.erreurs, self.avertissements)

    def _detect_columns(self, ws) -> Optional[Dict[str, int]]:
        """Détecte automatiquement les colonnes."""
        mapping = {}

        # Parcours des premières lignes
        for row_idx in range(1, min(15, ws.max_row + 1)):
            for col_idx in range(1, min(30, ws.max_column + 1)):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '').lower().strip()

                if not cell_value:
                    continue

                for field, pattern in self.PATTERNS_COLONNES.items():
                    if re.search(pattern, cell_value, re.IGNORECASE):
                        if field not in mapping:
                            mapping[field] = col_idx

        # La référence est obligatoire
        return mapping if 'reference' in mapping else None

    def _detect_header_and_charges(
        self, ws, mapping: Dict[str, int]
    ) -> Tuple[int, Dict[int, float]]:
        """Détecte la ligne d'en-tête et les colonnes de charges."""
        charges: Dict[int, int] = {}  # {col_idx: charge_value}
        header_row = 1

        # Recherche dans les premières lignes
        for row_idx in range(1, min(15, ws.max_row + 1)):
            row_charges: Dict[int, int] = {}

            for col_idx in range(1, min(50, ws.max_column + 1)):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '').strip()

                # Détection colonnes de charge
                # Patterns: "250", "300 kg/m²", "3.5 kN/m²", etc.
                match = re.match(r'^(\d{3,4})(\s*(kg|dan|kn))?', cell_value, re.IGNORECASE)
                if match:
                    charge = int(match.group(1))
                    if 150 <= charge <= 1200:
                        row_charges[col_idx] = charge

            # Si on trouve au moins 3 charges sur une ligne, c'est l'en-tête
            if len(row_charges) >= 3:
                charges = row_charges
                header_row = row_idx
                break

        return header_row, charges

    def _parse_row(
        self, ws, row_idx: int, mapping: Dict[str, int], charges: Dict[int, int]
    ) -> Optional[Dict[str, Any]]:
        """Parse une ligne de données."""
        ref_col = mapping.get('reference')
        ref_value = ws.cell(row=row_idx, column=ref_col).value if ref_col else None

        if not ref_value or str(ref_value).strip() == '':
            return None

        # Vérification que ce n'est pas une ligne d'en-tête ou de séparation
        ref_str = str(ref_value).strip().lower()
        if any(word in ref_str for word in ['type', 'référence', 'poutrelle', 'total']):
            return None

        # Extraction portées limites
        portees_limites = {}
        for col_idx, charge in charges.items():
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                try:
                    # Gestion des formats: "5.20", "5,20", "520" (cm)
                    portee_str = str(cell_value).replace(',', '.').strip()
                    portee = float(portee_str)

                    # Si la valeur semble en cm (> 15), convertir en m
                    if portee > 15:
                        portee = portee / 100

                    # Portée plausible entre 1 et 12 mètres
                    if 1.0 < portee < 12.0:
                        portees_limites[str(charge)] = round(portee, 2)
                except (ValueError, TypeError):
                    pass

        if not portees_limites:
            return None

        # Extraction hauteur hourdis
        hauteur_hourdis = self._extract_int(ws, row_idx, mapping.get('hourdis'))
        if not hauteur_hourdis:
            # Essayer de détecter dans la référence (ex: "BP 113-16" -> 16)
            ref_match = re.search(r'[-_\s](\d{2})(?:\s|$|cm)', str(ref_value))
            if ref_match:
                hauteur_hourdis = int(ref_match.group(1))

        # Valeur par défaut si non trouvée
        if not hauteur_hourdis or hauteur_hourdis not in [12, 16, 20, 25]:
            hauteur_hourdis = 16  # Défaut courant

        # Extraction entraxe
        entraxe = self._extract_int(ws, row_idx, mapping.get('entraxe'))
        if not entraxe or entraxe < 40 or entraxe > 80:
            entraxe = 60  # Entraxe standard

        return {
            'reference_poutrelle': str(ref_value).strip(),
            'hauteur_hourdis_cm': hauteur_hourdis,
            'entraxe_cm': entraxe,
            'epaisseur_table_cm': self._extract_float(ws, row_idx, mapping.get('table'), 5.0),
            'portees_limites': portees_limites
        }

    def _extract_int(self, ws, row: int, col: Optional[int], default: int = 0) -> int:
        """Extrait une valeur entière d'une cellule."""
        if not col:
            return default
        value = ws.cell(row=row, column=col).value
        if value is None:
            return default
        try:
            return int(float(str(value).replace(',', '.')))
        except (ValueError, TypeError):
            return default

    def _extract_float(self, ws, row: int, col: Optional[int], default: float = 0.0) -> float:
        """Extrait une valeur flottante d'une cellule."""
        if not col:
            return default
        value = ws.cell(row=row, column=col).value
        if value is None:
            return default
        try:
            return float(str(value).replace(',', '.'))
        except (ValueError, TypeError):
            return default
