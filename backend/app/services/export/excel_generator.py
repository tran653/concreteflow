"""
Générateur de fichiers Excel pour nomenclatures et quantitatifs.
Utilise openpyxl pour créer des classeurs Excel.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List, Optional
from datetime import datetime
from io import BytesIO


class ExcelGenerator:
    """Générateur de fichiers Excel pour ConcreteFlow."""

    def __init__(self):
        self.wb = None
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')

    def generate_nomenclature(
        self,
        elements: List[Dict[str, Any]],
        projet: Dict[str, Any] = None,
        include_summary: bool = True
    ) -> bytes:
        """
        Génère une nomenclature Excel des éléments.

        Args:
            elements: Liste des éléments avec leurs caractéristiques
            projet: Informations du projet
            include_summary: Inclure un résumé

        Returns:
            Contenu Excel en bytes
        """
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "Nomenclature"

        # En-tête du document
        row = 1
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "NOMENCLATURE DES ÉLÉMENTS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1

        if projet:
            ws[f'A{row}'] = f"Projet: {projet.get('name', 'N/A')}"
            row += 1
            ws[f'A{row}'] = f"Client: {projet.get('client_name', 'N/A')}"
            row += 1

        ws[f'A{row}'] = f"Date: {datetime.now().strftime('%d/%m/%Y')}"
        row += 2

        # En-têtes de colonnes
        headers = [
            "Réf.", "Type", "Longueur (mm)", "Largeur (mm)",
            "Hauteur (mm)", "Béton", "Ferraillage", "Poids (kg)"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_align

        row += 1

        # Données des éléments
        total_weight = 0
        total_volume = 0

        for elem in elements:
            dims = elem.get('dimensions', {})
            length = dims.get('length_mm', 0)
            width = dims.get('width_mm', 0)
            height = dims.get('height_mm', 200)

            # Calcul volume et poids
            volume = (length * width * height) / 1e9  # m³
            weight = volume * 2500  # kg (densité béton ~ 2500 kg/m³)
            total_volume += volume
            total_weight += weight

            data = [
                elem.get('reference', ''),
                self._get_type_label(elem.get('type', '')),
                length,
                width,
                height,
                elem.get('classe_beton', 'C30/37'),
                elem.get('ferraillage', {}).get('resume', 'Voir note'),
                round(weight, 1)
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col >= 3 and col <= 5:
                    cell.alignment = self.center_align
                elif col == 8:
                    cell.alignment = Alignment(horizontal='right')

            row += 1

        # Ligne de total
        row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=7, value=f"{total_volume:.2f} m³").font = Font(bold=True)
        ws.cell(row=row, column=8, value=f"{total_weight:.0f} kg").font = Font(bold=True)

        # Ajuster la largeur des colonnes
        column_widths = [10, 20, 15, 15, 15, 12, 25, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Résumé par type si demandé
        if include_summary:
            self._add_summary_sheet(elements)

        # Sauvegarder en bytes
        buffer = BytesIO()
        self.wb.save(buffer)
        content = buffer.getvalue()
        buffer.close()

        return content

    def generate_quantitatif(
        self,
        calculs: List[Dict[str, Any]],
        projet: Dict[str, Any] = None
    ) -> bytes:
        """
        Génère un quantitatif Excel des matériaux.

        Args:
            calculs: Liste des calculs avec résultats
            projet: Informations du projet

        Returns:
            Contenu Excel en bytes
        """
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "Quantitatif"

        # En-tête
        row = 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "QUANTITATIF MATÉRIAUX"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        if projet:
            ws[f'A{row}'] = f"Projet: {projet.get('name', 'N/A')}"
            row += 2

        # Section Béton
        ws[f'A{row}'] = "BÉTON"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        headers = ["Classe", "Volume (m³)", "Masse (kg)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        row += 1

        # Agréger par classe de béton
        beton_totals = {}
        for calcul in calculs:
            classe = calcul.get('parametres', {}).get('materiaux', {}).get('classe_beton', 'C30/37')
            geom = calcul.get('parametres', {}).get('geometrie', {})
            volume = geom.get('portee', 0) * geom.get('largeur', 0) * geom.get('hauteur', 0)

            if classe not in beton_totals:
                beton_totals[classe] = 0
            beton_totals[classe] += volume

        for classe, volume in beton_totals.items():
            ws.cell(row=row, column=1, value=classe).border = self.border
            ws.cell(row=row, column=2, value=round(volume, 2)).border = self.border
            ws.cell(row=row, column=3, value=round(volume * 2500, 0)).border = self.border
            row += 1

        row += 2

        # Section Acier
        ws[f'A{row}'] = "ACIER"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        headers = ["Type", "Diamètre", "Longueur (m)", "Masse (kg)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        row += 1

        # Agréger l'acier
        acier_totals = {}
        for calcul in calculs:
            ferraillage = calcul.get('resultats', {}).get('ferraillage', {})

            # Armatures inférieures
            inf = ferraillage.get('armatures_inferieures', {})
            if inf.get('designation'):
                key = ('Longitudinal', inf.get('designation'))
                if key not in acier_totals:
                    acier_totals[key] = {'longueur': 0, 'masse': 0}
                portee = calcul.get('parametres', {}).get('geometrie', {}).get('portee', 0)
                acier_totals[key]['longueur'] += portee

        for (type_acier, designation), totals in acier_totals.items():
            ws.cell(row=row, column=1, value=type_acier).border = self.border
            ws.cell(row=row, column=2, value=designation).border = self.border
            ws.cell(row=row, column=3, value=round(totals['longueur'], 1)).border = self.border
            ws.cell(row=row, column=4, value="-").border = self.border
            row += 1

        # Ajuster colonnes
        for i, width in enumerate([15, 15, 15, 15], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Sauvegarder
        buffer = BytesIO()
        self.wb.save(buffer)
        content = buffer.getvalue()
        buffer.close()

        return content

    def generate_plan_pose(
        self,
        elements: List[Dict[str, Any]],
        plan_info: Dict[str, Any] = None
    ) -> bytes:
        """
        Génère un plan de pose simplifié en Excel.

        Args:
            elements: Liste des éléments avec positions
            plan_info: Informations du plan

        Returns:
            Contenu Excel en bytes
        """
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "Plan de pose"

        # En-tête
        row = 1
        ws[f'A{row}'] = "PLAN DE POSE"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2

        # En-têtes
        headers = ["Réf.", "X (mm)", "Y (mm)", "Rotation (°)", "Longueur", "Largeur"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        row += 1

        # Données
        for elem in elements:
            pos = elem.get('position', {})
            dims = elem.get('dimensions', {})

            data = [
                elem.get('reference', ''),
                pos.get('x', 0),
                pos.get('y', 0),
                pos.get('rotation', 0),
                dims.get('length_mm', 0),
                dims.get('width_mm', 0)
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                cell.alignment = self.center_align
            row += 1

        # Ajuster colonnes
        for i, width in enumerate([10, 12, 12, 12, 12, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        buffer = BytesIO()
        self.wb.save(buffer)
        content = buffer.getvalue()
        buffer.close()

        return content

    def _add_summary_sheet(self, elements: List[Dict]):
        """Ajoute une feuille de résumé."""
        ws = self.wb.create_sheet("Résumé")

        # Compter par type
        type_counts = {}
        for elem in elements:
            elem_type = elem.get('type', 'autre')
            if elem_type not in type_counts:
                type_counts[elem_type] = {'count': 0, 'total_length': 0}
            type_counts[elem_type]['count'] += 1
            type_counts[elem_type]['total_length'] += elem.get('dimensions', {}).get('length_mm', 0)

        row = 1
        ws[f'A{row}'] = "RÉSUMÉ PAR TYPE"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2

        headers = ["Type", "Quantité", "Longueur totale (m)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        row += 1

        for type_name, data in type_counts.items():
            ws.cell(row=row, column=1, value=self._get_type_label(type_name)).border = self.border
            ws.cell(row=row, column=2, value=data['count']).border = self.border
            ws.cell(row=row, column=3, value=round(data['total_length'] / 1000, 1)).border = self.border
            row += 1

        for i, width in enumerate([20, 12, 18], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _get_type_label(self, type_produit: str) -> str:
        """Retourne le libellé du type de produit."""
        labels = {
            'poutrelle': 'Poutrelle',
            'predalle': 'Prédalle',
            'dalle_alveolaire': 'Dalle alvéolaire',
            'poutre': 'Poutre',
            'dalle_pleine': 'Dalle pleine',
        }
        return labels.get(type_produit, type_produit)


def generate_nomenclature(
    elements: List[Dict[str, Any]],
    projet: Dict[str, Any] = None
) -> bytes:
    """
    Fonction utilitaire pour générer une nomenclature Excel.
    """
    generator = ExcelGenerator()
    return generator.generate_nomenclature(elements, projet)
